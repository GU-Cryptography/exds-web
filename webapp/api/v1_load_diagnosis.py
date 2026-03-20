# -*- coding: utf-8 -*-
"""
负荷数据校核模块 API
提供负荷数据完整性校核、误差分析、数据聚合等功能
"""

import logging
from datetime import datetime, timedelta
from typing import List, Optional, Dict
from bson import ObjectId
from fastapi import APIRouter, Query, HTTPException, Depends, File, UploadFile, Form
from webapp.tools.mongo import DATABASE
from webapp.tools.security import get_current_active_user, User
from webapp.services.meter_data_import_service import MeterDataImportService
from webapp.services.mp_data_import_service import MpDataImportService
from webapp.services.load_aggregation_service import LoadAggregationService
from webapp.api.dependencies.authz import require_permission

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/load-data", tags=["load-data"])

# 集合定义
CUSTOMER_ARCHIVES = DATABASE['customer_archives']
UNIFIED_LOAD_CURVE = DATABASE['unified_load_curve']
RAW_MP_DATA = DATABASE['raw_mp_data']
RAW_METER_DATA = DATABASE['raw_meter_data']
RETAIL_CONTRACTS = DATABASE['retail_contracts']

from webapp.models.load_enums import FusionStrategy
from webapp.services.load_query_service import LoadQueryService
from webapp.services.diagnosis_service import DiagnosisService


# ##############################################################################
# 主页面接口
# ##############################################################################




@router.get("/customers", summary="获取客户校核列表")
async def get_load_data_customers(
    status: Optional[str] = Query(None, description="筛选状态: anomaly/error/pending"),
    search: Optional[str] = Query(None, description="搜索客户名称"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(10, ge=1, le=100, description="每页大小"),
    current_user: User = Depends(get_current_active_user)
):
    """
    获取客户列表 (v3.0)
    返回字段:
    - integrity_rate: 完整率
    - reliability_rate: 可靠率
    - accuracy_rate: 准确率
    - contract_days: 合同天数 (估算)
    - status: 状态 ('normal', 'pending', 'warning', 'critical')
    """
    try:
        from datetime import datetime as dt
        today = dt.now()

        # 1. 签约客户
        signed_contracts = RETAIL_CONTRACTS.find({
            "purchase_start_month": {"$lte": today},
            "purchase_end_month": {"$gte": today}
        })
        
        signed_customers = {} # cid -> {name, start, end}
        for c in signed_contracts:
            cid = c.get("customer_id")
            cname = c.get("customer_name")
            if cid and cname:
                signed_customers[cid] = {
                    "name": cname,
                    "start": c.get("purchase_start_month"),
                    "end": c.get("purchase_end_month")
                }
                
        customer_ids = list(signed_customers.keys())

        # 2. 聚合统计
        pipeline = [
            {"$match": {"customer_id": {"$in": customer_ids}}},
            {"$addFields": {
                "has_missing_mp": {
                    "$gt": [{"$size": {"$ifNull": ["$mp_load.missing_mps", []]}}, 0]
                }
            }},
            {"$group": {
                "_id": "$customer_id",
                "mp_dates": {"$push": {"$cond": [{"$ne": ["$mp_load", None]}, "$date", "$$REMOVE"]}},
                "meter_dates": {"$push": {"$cond": [{"$ne": ["$meter_load", None]}, "$date", "$$REMOVE"]}},
                "all_dates": {"$push": "$date"},
                "missing_mp_days": {"$sum": {"$cond": ["$has_missing_mp", 1, 0]}},
                "deviations": {"$push": "$deviation"}
            }}
        ]
        
        load_stats = {
            doc["_id"]: doc 
            for doc in UNIFIED_LOAD_CURVE.aggregate(pipeline)
        }
        
        result = []
        for cid, info in signed_customers.items():
            if search and search.lower() not in info['name'].lower():
                continue
            
            stats = load_stats.get(cid, {})
            
            # --- Metrics Calculation ---
            
            # 1. Integrity (完整率) & Cycle (周期/天数)
            # User Definition:
            # Cycle: Earliest record date ~ Latest record date
            # Days: Count of records in unified_load_curve
            # Integrity Rate: Record Count / (Yesterday - Earliest Record Date)
            
            unique_dates = sorted(list(set(stats.get("all_dates", []))))
            record_count = len(unique_dates)
            
            cycle_range = "-"
            data_days = 0 
            integrity_rate = 0.0

            if unique_dates:
                # Cycle display
                min_date_str = unique_dates[0]
                max_date_str = unique_dates[-1]
                cycle_range = f"{min_date_str}~{max_date_str}"
                data_days = record_count
                
                # Integrity Calculation
                try:
                    min_date = dt.strptime(min_date_str, "%Y-%m-%d")
                    yesterday = today.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
                    
                    # Denominator: Days from Earliest Record to Yesterday
                    expected_days = (yesterday - min_date).days + 1
                    
                    if expected_days > 0:
                        integrity_rate = record_count / expected_days
                    else:
                        integrity_rate = 1.0 if record_count > 0 else 0.0

                except Exception:
                    integrity_rate = 0.0
            
            integrity_rate = min(1.0, integrity_rate)
            
            # 2. Reliability (可靠率/异常天数)
            missing_mp_days = stats.get("missing_mp_days", 0)
            
            # 3. Accuracy (准确率)
            daily_errors = []
            has_warning = False
            for dev in stats.get("deviations", []):
                if dev and isinstance(dev, dict):
                    if dev.get("daily_error") is not None:
                        daily_errors.append(abs(dev["daily_error"]))
                    if dev.get("is_warning"):
                        has_warning = True
            
            avg_error = sum(daily_errors) / len(daily_errors) if daily_errors else 0.0
            accuracy_rate = 1.0 - avg_error
            accuracy_rate = max(0.0, accuracy_rate)

            # Logic: 
            # Critical: Integrity < 80% OR Accuracy < 95% (Error > 5%)
            # Warning: Reliability issue (missing_mp_days > 0)
            # Pending: Has one side missing? (Simplified check)
            # Normal: Else
            
            mp_count = len(stats.get("mp_dates", []))
            meter_count = len(stats.get("meter_dates", []))
            is_pending = mp_count != meter_count
            
            status_code = 'normal'
            if integrity_rate < 0.8 or accuracy_rate < 0.95:
                status_code = 'critical'
            elif missing_mp_days > 0:
                status_code = 'warning'
            elif is_pending:
                status_code = 'pending'

            # Filter by status if requested
            if status:
                map_status = {
                    'anomaly': ['critical', 'warning'], # Map 'abnormal/limit' to critical/warning
                    'error': ['critical'],
                    'pending': ['pending']
                }
                # Mapping user's query param to our new status
                # If status is 'anomaly' (integrity), check integrity
                # If status is 'error' (accuracy), check accuracy
                # If status is 'pending', check pending
                
                # To be precise with legacy params:
                match = False
                if status == 'anomaly' and (integrity_rate < 0.9 or missing_mp_days > 0): match = True
                if status == 'error' and accuracy_rate < 0.95: match = True
                if status == 'pending' and is_pending: match = True
                if status == 'reliability' and missing_mp_days > 0: match = True
                if status == 'pending_meter' and is_pending: match = True  # Uses same logic as pending for now
                if not match:
                    continue

            result.append({
                "customer_id": cid,
                "customer_name": info['name'],
                "cycle_range": cycle_range,
                "data_days": data_days,
                "integrity_rate": integrity_rate,
                "reliability_issue_days": missing_mp_days,
                "accuracy_rate": accuracy_rate,
                "data_distribution": {
                    "mp": mp_count,
                    "meter": meter_count
                },
                "status": status_code
            })

        # Sorting & Pagination
        # Sort by status priority (Critical > Warning > Pending > Normal)
        status_priority = {'critical': 0, 'warning': 1, 'pending': 2, 'normal': 3}
        result.sort(key=lambda x: status_priority.get(x['status'], 3))
        
        total = len(result)
        start = (page - 1) * page_size
        end = start + page_size
        paged_result = result[start:end]
        
        return {
            "total": total,
            "customers": paged_result
        }
    except Exception as e:
        logger.error(f"获取客户列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


def _calculate_continuity(dates: List[str]) -> float:
    """计算日期连续性（无缺天的比例）"""
    if not dates or len(dates) < 2:
        return 1.0 if dates else 0.0
    
    start = datetime.strptime(dates[0], "%Y-%m-%d")
    end = datetime.strptime(dates[-1], "%Y-%m-%d")
    expected_days = (end - start).days + 1
    actual_days = len(set(dates))
    
    return actual_days / expected_days if expected_days > 0 else 0.0


# ##############################################################################
# 客户详情接口
# ##############################################################################

@router.get("/customers/{customer_id}", summary="获取客户详情")
async def get_customer_detail(
    customer_id: str,
    current_user: User = Depends(get_current_active_user)
):
    """
    获取客户详情，包含：
    - 基础统计（户号数/电表数/计量点数）
    - 质量指标（断点天数/缺失天数/准确率）
    - 电表档案
    """
    try:
        from bson import ObjectId
        
        # 获取客户档案
        customer = CUSTOMER_ARCHIVES.find_one({"_id": customer_id})
        if not customer:
            customer = CUSTOMER_ARCHIVES.find_one({"_id": ObjectId(customer_id)})
        
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在")
        
        # 统计户号数/电表数/计量点数
        account_count = len(customer.get("accounts", []))
        meter_count = 0
        mp_count = 0
        accounts = []
        
        for account in customer.get("accounts", []):
            account_meters = account.get("meters", [])
            account_mps = account.get("metering_points", [])
            meter_count += len(account_meters)
            mp_count += len(account_mps)
            
            accounts.append({
                "account_no": account.get("account_no", ""),
                "meters": [{
                    "meter_id": m.get("meter_id"),
                    "multiplier": m.get("multiplier", 1),
                    "allocation_ratio": m.get("allocation_ratio")
                } for m in account_meters],
                "metering_points": [mp.get("mp_no") for mp in account_mps]
            })
        # 获取该客户的所有负荷曲线数据（用于计算质量指标）
        curves = list(UNIFIED_LOAD_CURVE.find(
            {"customer_id": customer_id}
        ).sort("date", -1))
        
        # 计算数据概览
        all_dates = sorted(set(c["date"] for c in curves))
        mp_dates = sorted([c["date"] for c in curves if c.get("mp_load")])
        meter_dates = sorted([c["date"] for c in curves if c.get("meter_load")])
        
        # 计算断点天数（期望天数 - 实际天数）
        if all_dates:
            start = datetime.strptime(all_dates[0], "%Y-%m-%d")
            end = datetime.strptime(all_dates[-1], "%Y-%m-%d")
            expected_days = (end - start).days + 1
            actual_days = len(all_dates)
            gap_days = expected_days - actual_days
        else:
            gap_days = 0
            expected_days = 0
            actual_days = 0
        
        # 获取客户当前合同（当日在合同周期内）
        RETAIL_CONTRACTS = DATABASE['retail_contracts']
        today = datetime.now()
        contract = RETAIL_CONTRACTS.find_one({
            "customer_id": customer_id,
            "purchase_start_month": {"$lte": today},
            "purchase_end_month": {"$gte": today}
        })
        
        # 如果没有当前合同，则取最新的合同
        if not contract:
            contract = RETAIL_CONTRACTS.find_one(
                {"customer_id": customer_id},
                sort=[("purchase_start_month", -1)]  # 取最新的合同
            )
        
        # 确定MP期望日期范围：合同开始日期 → (当日-2)
        mp_expected_end = today - timedelta(days=2)  # MP数据有2天延迟
        
        if contract and contract.get("purchase_start_month"):
            mp_expected_start = contract["purchase_start_month"]
            if isinstance(mp_expected_start, str):
                mp_expected_start = datetime.strptime(mp_expected_start, "%Y-%m-%d")
        else:
            # 无合同则使用第一条数据日期作为起始
            mp_expected_start = datetime.strptime(all_dates[0], "%Y-%m-%d") if all_dates else mp_expected_end
        
        # 生成期望的MP日期集合
        expected_mp_dates = set()
        curr = mp_expected_start
        while curr <= mp_expected_end:
            expected_mp_dates.add(curr.strftime("%Y-%m-%d"))
            curr += timedelta(days=1)
        
        # 实际有完整MP数据的日期（actual == expected）
        # 实际有部分MP数据的日期（0 < actual < expected）也算缺失
        complete_mp_dates = set()
        partial_mp_dates = set()
        
        for c in curves:
            date = c.get("date")
            if c.get("mp_load"):
                mp_actual = c["mp_load"].get("mp_count", 0)
                if mp_actual > 0:
                    if mp_actual >= mp_count:
                        complete_mp_dates.add(date)
                    else:
                        partial_mp_dates.add(date)  # 部分缺失
        
        # MP缺失天数 = 
        #   1. 期望日期中完全没有MP数据的天数
        #   2. 加上有数据但部分缺失的天数
        missing_mp_dates = expected_mp_dates - complete_mp_dates - partial_mp_dates
        mp_incomplete_days = len(missing_mp_dates) + len(partial_mp_dates)
        
        # 计算电表缺失天数和最大误差（仍基于现有记录）
        meter_incomplete_days = 0
        max_error = 0.0
        
        for c in curves:
            # Meter 缺失
            if c.get("meter_load"):
                meter_l = c["meter_load"]
                # 修正逻辑：不再对比 meter_count < total_meter_count (因为档案可能变更)
                # 而是检查聚合时记录的 missing_meters 是否存在且非空
                if meter_l.get("missing_meters"):
                    meter_incomplete_days += 1
            
            # 最大误差
            if c.get("deviation"):
                err = c["deviation"].get("daily_error")
                if err is not None:
                    # 2026-01-29 更新：仅统计 2026-01-01 及之后的误差
                    is_valid_date = True
                    date_val = c.get("date")
                    if date_val:
                        try:
                            if date_val < "2026-01-01":
                                is_valid_date = False
                        except:
                            pass
                    
                    if is_valid_date:
                        abs_err = abs(err)
                        if abs_err > max_error:
                            max_error = abs_err
        
        # 获取日期范围
        date_range = f"{all_dates[0]} ~ {all_dates[-1]}" if all_dates else None
        
        return {
            "customer_id": customer_id,
            "customer_name": customer.get("user_name", "未知"),
            "stats": {
                "account_count": account_count,
                "meter_count": meter_count,
                "mp_count": mp_count
            },
            "quality": {
                "gap_days": gap_days,  # 无数据
                "mp_incomplete_days": mp_incomplete_days,  # 计量点缺失
                "meter_incomplete_days": meter_incomplete_days,  # 电表缺失
                "max_error": round(max_error * 100, 1),  # 误差 (%)
                "total_days": actual_days
            },
            "date_range": date_range,
            "accounts": accounts
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取客户详情失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取客户详情失败: {str(e)}")


@router.get("/customers/{customer_id}/calendar", summary="获取热力图日历数据")
async def get_customer_calendar(
    customer_id: str,
    month: Optional[str] = Query(None, description="月份, 格式 YYYY-MM"),
    fetch_all: bool = Query(False, alias="all", description="是否获取全部数据（用于时间线视图），忽略 month 参数"),
    current_user: User = Depends(get_current_active_user)
):
    """
    获取热力图日历数据，显示每天的采集完整度
    """
    try:
        # 确定查询日期范围
        if fetch_all:
            # 获取全部数据（不限制日期范围）
            start_date = None
            end_date = None
        elif month:
            year, mon = map(int, month.split("-"))
            start_date = f"{year:04d}-{mon:02d}-01"
            if mon == 12:
                end_date = f"{year+1:04d}-01-01"
            else:
                end_date = f"{year:04d}-{mon+1:02d}-01"
        else:
            # 默认最近30天
            today = datetime.now()
            start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
            end_date = (today + timedelta(days=1)).strftime("%Y-%m-%d")
        
        # 先查询客户档案获取总计量点数和总电表数
        customer = CUSTOMER_ARCHIVES.find_one({"_id": customer_id})
        if not customer:
            from bson import ObjectId
            try:
                customer = CUSTOMER_ARCHIVES.find_one({"_id": ObjectId(customer_id)})
            except:
                pass
        
        # 统计档案中的总计量点数和总电表数
        total_mp_count = 0
        total_meter_count = 0
        if customer:
            for account in customer.get("accounts", []):
                total_mp_count += len(account.get("metering_points", []))
                total_meter_count += len(account.get("meters", []))
        
        # 查询曲线数据
        query = {"customer_id": customer_id}
        if start_date and end_date:
            query["date"] = {"$gte": start_date, "$lt": end_date}
        
        curves = list(UNIFIED_LOAD_CURVE.find(
            query,
            {"date": 1, "mp_load": 1, "meter_load": 1, "deviation": 1}
        ).sort("date", 1))
        
        # 构建日历数据
        calendar_data = []
        for curve in curves:
            date = curve.get("date")
            mp_load = curve.get("mp_load", {})
            meter_load = curve.get("meter_load", {})
            
            # 计量点数据：实际值从曲线取，总数从档案取
            mp_actual = mp_load.get("mp_count", 0) if mp_load else 0
            mp_expected = total_mp_count
            
            # 电表数据：实际值从曲线取，总数从档案取
            meter_actual = meter_load.get("meter_count", 0) if meter_load else 0
            meter_expected = total_meter_count
            
            # 计算当日误差
            deviation = curve.get("deviation", {})
            daily_error = deviation.get("daily_error") if deviation else None
            
            # 判断是否有数据
            has_mp = mp_actual > 0
            has_meter = meter_actual > 0
            
            calendar_data.append({
                "date": date,
                "has_mp_data": has_mp,
                "has_meter_data": has_meter,
                "daily_error": (
                    round(daily_error * 100, 2) 
                    if daily_error is not None and abs(daily_error) != float('inf') and daily_error == daily_error
                    else None
                ),
                "mp_actual": mp_actual,
                "mp_expected": mp_expected,
                "meter_actual": meter_actual,
                "meter_expected": meter_expected,
                "missing_mps": mp_load.get("missing_mps", []) if mp_load else [],
                "missing_meters": meter_load.get("missing_meters", []) if meter_load else []
            })
        
        # 获取缺失日期列表（仅在指定日期范围时计算）
        missing_dates = []
        if start_date and end_date:
            all_dates = set()
            current = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            while current < end:
                all_dates.add(current.strftime("%Y-%m-%d"))
                current += timedelta(days=1)
            
            existing_dates = {c["date"] for c in calendar_data}
            missing_dates = sorted(all_dates - existing_dates)
        
        # 不完整日期：计量点或电表数据不完整
        incomplete_dates = [c["date"] for c in calendar_data 
                          if c["mp_actual"] < c["mp_expected"] or c["meter_actual"] < c["meter_expected"]]
        
        return {
            "customer_id": customer_id,
            "total_mp_count": total_mp_count,
            "total_meter_count": total_meter_count,
            "calendar": calendar_data,
            "missing_dates": missing_dates,
            "incomplete_dates": incomplete_dates
        }
    except Exception as e:
        logger.error(f"获取日历数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取日历数据失败: {str(e)}")


@router.get("/customers/{customer_id}/curves", summary="获取曲线对比数据")
async def get_customer_curves(
    customer_id: str,
    start_date: str = Query(..., description="开始日期, 格式 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期, 格式 YYYY-MM-DD"),
    detail_date: Optional[str] = Query(None, description="详情日期（返回48点数据）"),
    current_user: User = Depends(get_current_active_user)
):
    """
    获取曲线对比数据：
    - 如果指定 detail_date，返回该日期的48点曲线对比
    - 否则返回日期范围内的日电量对比
    """
    try:
        return LoadQueryService.get_diagnosis_curves(
            customer_id=customer_id,
            start_date=start_date,
            end_date=end_date,
            detail_date=detail_date
        )
    except Exception as e:
        logger.error(f"获取曲线对比数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取曲线对比数据失败: {str(e)}")


# ##############################################################################
# 聚合与重算接口
# ##############################################################################

@router.post("/reaggregate", summary="触发重新聚合")
async def trigger_reaggregate(
    data_type: str = Query("all", description="数据类型: mp/meter/all"),
    customer_id: Optional[str] = Query(None, description="指定客户ID，为空则处理所有客户"),
    start_date: Optional[str] = Query(None, description="开始日期（可选，默认自动检测）"),
    end_date: Optional[str] = Query(None, description="结束日期（可选，默认自动检测）"),
    mode: str = Query("incremental", description="聚合模式: incremental(增量)/full(全量)"),
    delete_existing: bool = Query(False, description="是否删除原数据（仅全量模式有效）"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_monthly_manual_import:edit"))
):
    """
    触发重新聚合计算（真正的增量逻辑）
    
    业务逻辑：
    1. 获取原始数据中所有 (customer_id, date) 对
    2. 与 unified_load_curve 对比，找出尚未聚合的日期
    3. 只处理缺失的日期（真正的增量）
    """
    import time
    start_time = time.time()
    
    try:
        from datetime import datetime as dt
        from bson import ObjectId
        
        # 确定客户列表
        if customer_id:
            customer_ids = [customer_id]
        else:
            # 获取所有签约客户
            today_dt = dt.now()
            signed_contracts = RETAIL_CONTRACTS.find({
                "purchase_start_month": {"$lte": today_dt},
                "purchase_end_month": {"$gte": today_dt}
            })
            customer_ids = list(set(
                c.get("customer_id") for c in signed_contracts if c.get("customer_id")
            ))
        
        if not customer_ids:
            return {
                "status": "completed",
                "message": "没有找到需要处理的客户",
                "processed": 0,
                "updated": 0
            }
        
        # 获取客户档案信息（计量点和电表）
        customer_info = {}  # customer_id -> {name, mp_ids, meter_ids}
        for cust in CUSTOMER_ARCHIVES.find({"_id": {"$in": [ObjectId(cid) if len(cid) == 24 else cid for cid in customer_ids]}}):
            cid = str(cust.get("_id"))
            mp_list = []
            meter_list = []
            for account in cust.get("accounts", []):
                for mp in account.get("metering_points", []):
                    if mp.get("mp_no"):
                        mp_list.append(mp["mp_no"])
                for meter in account.get("meters", []):
                    if meter.get("meter_id"):
                        meter_list.append(meter["meter_id"])
            customer_info[cid] = {
                "name": cust.get("user_name", "未知"),
                "mp_ids": mp_list,
                "meter_ids": meter_list
            }
        
        # 收集所有需要的 mp_ids 和 meter_ids
        # 仅在全量模式下需要手动查询原始数据日期 (增量模式由 get_pending_tasks 内部处理)
        raw_mp_dates = {}
        raw_meter_dates = {}
        aggregated_status = {}
        
        if mode == "full":
            all_mp_ids = [mp for info in customer_info.values() for mp in info["mp_ids"]]
            all_meter_ids = [m for info in customer_info.values() for m in info["meter_ids"]]
            
            # 查询原始数据的日期
            if all_mp_ids:
                for doc in RAW_MP_DATA.find({"mp_id": {"$in": all_mp_ids}}, {"mp_id": 1, "date": 1}):
                    mp_id = doc.get("mp_id")
                    date = doc.get("date")
                    if mp_id and date:
                        raw_mp_dates.setdefault(mp_id, set()).add(date)
            
            if all_meter_ids:
                for doc in RAW_METER_DATA.find({"meter_id": {"$in": all_meter_ids}}, {"meter_id": 1, "date": 1}):
                    meter_id = doc.get("meter_id")
                    date = doc.get("date")
                    if meter_id and date:
                        raw_meter_dates.setdefault(meter_id, set()).add(date)
        
        # 确定每个客户需要处理的日期
        customer_pending_dates = {}  # customer_id -> set of pending dates

        # 如果是全量重算模式 (full mode)
        if mode == "full":
            # 2026-01-29 Feature: 支持删除原数据
            if delete_existing:
                del_query = {"customer_id": {"$in": list(customer_info.keys())}}
                # 如果指定了日期范围，只删除范围内的
                if start_date or end_date:
                    date_filter = {}
                    if start_date: date_filter["$gte"] = start_date
                    if end_date: date_filter["$lte"] = end_date
                    del_query["date"] = date_filter
                
                del_result = UNIFIED_LOAD_CURVE.delete_many(del_query)
                logger.info(f"全量重算: 已删除 {del_result.deleted_count} 条旧数据 (customers={len(customer_info)})")

            # 在全量模式下，如果没有指定 delete_existing，通常我们也希望重新计算所有涉及的日期
            # 逻辑：获取所有原始数据的日期并加入 pending
            for cid, info in customer_info.items():
                current_raw_mp_dates = set()
                for mp_id in info["mp_ids"]:
                    current_raw_mp_dates.update(raw_mp_dates.get(mp_id, set()))
                
                current_raw_meter_dates = set()
                for meter_id in info["meter_ids"]:
                    current_raw_meter_dates.update(raw_meter_dates.get(meter_id, set()))
                
                all_dates = current_raw_mp_dates | current_raw_meter_dates
                
                # 过滤日期范围
                if start_date:
                    all_dates = {d for d in all_dates if d >= start_date}
                if end_date:
                    all_dates = {d for d in all_dates if d <= end_date}
                
                if all_dates:
                    customer_pending_dates[cid] = sorted(all_dates)

        else:
            # 增量模式 (incremental) - 使用新的服务方法，支持新鲜度检查
            try:
                customer_pending_dates = LoadAggregationService.get_pending_tasks(
                    customer_ids=customer_ids,
                    start_date=start_date,
                    end_date=end_date
                )
            except Exception as e:
                logger.error(f"获取增量聚合任务失败: {e}")
                customer_pending_dates = {}
        
        # 统计要处理的总数
        total_pending = sum(len(dates) for dates in customer_pending_dates.values())
        
        if total_pending == 0:
            elapsed = time.time() - start_time
            # 如果是全量删除模式且删除了数据但没新数据生成（罕见），也应该算完成
            msg = "所有数据已是最新，无需聚合"
            if mode == "full" and delete_existing:
                msg = "旧数据已删除，但无需生成新数据（无原始数据）"

            return {
                "status": "completed",
                "message": msg,
                "customer_count": len(customer_ids),
                "processed": 0,
                "updated": 0,
                "elapsed_seconds": round(elapsed, 2)
            }
        
        # 处理统计
        processed = 0
        updated = 0
        errors = []
        
        for cid, pending_dates in customer_pending_dates.items():
            customer_name = customer_info.get(cid, {}).get("name", "未知")
            
            for date in pending_dates:
                try:
                    result = LoadAggregationService.upsert_unified_load_curve(
                        cid, date, customer_name
                    )
                    if result:
                        updated += 1
                    processed += 1
                except Exception as e:
                    errors.append(f"{cid}@{date}: {str(e)}")
        
        elapsed = time.time() - start_time
        
        return {
            "status": "completed",
            "message": f"增量聚合完成",
            "data_type": data_type,
            "customer_count": len(customer_pending_dates),
            "total_pending": total_pending,
            "processed": processed,
            "updated": updated,
            "elapsed_seconds": round(elapsed, 2),
            "errors": errors[:10] if errors else []
        }
    except Exception as e:
        logger.error(f"触发重新聚合失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"触发重新聚合失败: {str(e)}")





# ##############################################################################
# 数据导入接口
# ##############################################################################

@router.post("/import/meter", summary="导入电表示度数据")
async def import_meter_data(
    file: UploadFile = File(..., description="Excel 文件"),
    overwrite: bool = Form(False, description="覆盖已存在数据"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_monthly_manual_import:edit"))
):
    """
    导入电表示度数据（Excel 文件）
    
    业务逻辑：
    1. 从文件名提取电表号
    2. 解析 Excel 内容（支持96点/1440点格式）
    3. 入库（根据 overwrite 决定是否覆盖）
    """
    try:
        # 读取文件内容
        content = await file.read()
        filename = file.filename
        
        # 调用导入服务
        result = MeterDataImportService.import_excel_file(content, filename, overwrite)
        
        return result
    except Exception as e:
        logger.error(f"导入电表数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")





@router.post("/import/mp", summary="导入计量点负荷数据")
async def import_mp_data(
    file: UploadFile = File(..., description="Excel 文件"),
    overwrite: bool = Form(False, description="覆盖已存在数据"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_monthly_manual_import:edit"))
):
    """
    导入计量点负荷数据（Excel 文件）
    
    业务逻辑：
    1. 解析 Excel 内容（支持24/48时段格式）
    2. 提取计量点ID、日期、时段电量
    3. 入库（根据 overwrite 决定是否覆盖）
    """
    try:
        content = await file.read()
        filename = file.filename
        
        result = MpDataImportService.import_excel_file(content, filename, overwrite)
        
        return result
    except Exception as e:
        logger.error(f"导入计量点数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")





# ##############################################################################
# 负荷曲线查询接口
# ##############################################################################




# ##############################################################################
# 系数校核接口
# ##############################################################################

@router.post("/calibration/preview", summary="预览校核状态")
async def preview_calibration(
    customer_id: str = Query(..., description="客户ID"),
    start_date: str = Query(..., description="开始日期 (YYYY-MM-DD)"),
    end_date: str = Query(..., description="结束日期 (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_monthly_manual_import:edit"))
):
    """
    预览各个户号的校验状态 (平衡/偏差/缺数)
    """
    try:
        from webapp.services.calibration_service import CalibrationService
        return CalibrationService.preview_calibration_status(customer_id, start_date, end_date)
    except Exception as e:
        logger.error(f"预览校核状态失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/calibration/calculate", summary="计算推荐系数 (最小二乘法)")
async def calculate_calibration_coefficients(
    customer_id: str = Query(..., description="客户ID"),
    start_date: str = Query(..., description="开始日期 (YYYY-MM-DD)"),
    end_date: str = Query(..., description="结束日期 (YYYY-MM-DD)"),
    account_no: Optional[str] = Query(None, description="指定户号 (可选)"),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_monthly_manual_import:edit"))
):
    """
    计算推荐的电表分配系数 (基于约束最小二乘法)
    如果指定了 account_no，则只计算该户号下的电表系数
    """
    try:
        from webapp.services.calibration_service import CalibrationService
        return CalibrationService.calculate_recommended_coefficients(customer_id, start_date, end_date, account_no)
    except Exception as e:
        logger.error(f"计算推荐系数失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/calibration/apply", summary="应用推荐系数")
async def apply_calibration_coefficients(
    request: dict, # {customer_id, coefficients, update_history, history_range}
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_monthly_manual_import:edit"))
):
    """
    应用推荐系数，并可选触发历史重算
    request body:
    {
        "customer_id": str,
        "coefficients": [{"meter_id": str, "value": float}, ...],
        "update_history": bool,
        "history_range": [start_date, end_date] (optional)
    }
    """
    try:
        from webapp.services.calibration_service import CalibrationService
        customer_id = request.get("customer_id")
        coefficients = request.get("coefficients", [])
        update_history = request.get("update_history", False)
        history_range = request.get("history_range")
        
        if not customer_id or not coefficients:
            raise HTTPException(status_code=400, detail="Missing required parameters")
            
        return CalibrationService.apply_coefficients(
            customer_id, coefficients, update_history, history_range
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"应用系数失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ##############################################################################
# 诊断接口
# ##############################################################################

@router.get("/signed-customers", summary="获取签约客户列表")
async def get_signed_customers(
    current_user: User = Depends(get_current_active_user)
):
    """
    获取当前所有签约客户列表（用于诊断初始化）
    只返回客户ID和名称，不返回诊断数据
    """
    try:
        customers = DiagnosisService.get_signed_customers()
        return {
            "total": len(customers),
            "customers": customers
        }
    except Exception as e:
        logger.error(f"获取签约客户列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/diagnose", summary="执行批量诊断")
async def diagnose_all_customers(
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_load_validation:edit"))
):
    """
    诊断所有签约客户的数据质量
    
    返回:
    - summary: 统计摘要（各类问题客户数）
    - customers: 详细诊断结果列表
    """
    try:
        result = DiagnosisService.diagnose_all_customers()
        return result
    except Exception as e:
        logger.error(f"执行诊断失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))





# ##############################################################################
# 计量点缺失导出接口
# ##############################################################################

from fastapi.responses import StreamingResponse
import io
from urllib.parse import quote

@router.get("/export/mp-missing", summary="导出计量点缺失明细")
async def export_mp_missing(
    current_user: User = Depends(get_current_active_user)
):
    """
    导出所有签约客户的计量点缺失明细Excel
    
    表格格式：
    - 客户名称
    - 计量点号
    - 缺失日期（逗号分隔）
    - 缺失天数
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 获取所有签约客户
        today = datetime.now()
        signed_customers = list(RETAIL_CONTRACTS.find({
            "purchase_start_month": {"$lte": today},
            "purchase_end_month": {"$gte": today}
        }))
        
        customer_ids = list(set(c.get("customer_id") for c in signed_customers if c.get("customer_id")))
        
        # 收集所有缺失数据
        rows = []
        
        for customer_id in customer_ids:
            # 获取客户档案
            customer = CUSTOMER_ARCHIVES.find_one({"_id": customer_id})
            if not customer:
                try:
                    from bson import ObjectId
                    customer = CUSTOMER_ARCHIVES.find_one({"_id": ObjectId(customer_id)})
                except:
                    pass
            
            if not customer:
                continue
                
            customer_name = customer.get("user_name", customer_id)
            
            # 获取所有计量点
            mp_list = []
            for account in customer.get("accounts", []):
                for mp in account.get("metering_points", []):
                    mp_no = mp.get("mp_no")
                    if mp_no:
                        mp_list.append(mp_no)
            
            if not mp_list:
                continue
            
            # 获取合同周期
            contract = RETAIL_CONTRACTS.find_one({
                "customer_id": customer_id,
                "purchase_start_month": {"$lte": today},
                "purchase_end_month": {"$gte": today}
            })
            
            if not contract:
                contract = RETAIL_CONTRACTS.find_one(
                    {"customer_id": customer_id},
                    sort=[("purchase_start_month", -1)]
                )
            
            if not contract:
                continue
            
            mp_expected_start = contract.get("purchase_start_month")
            if isinstance(mp_expected_start, str):
                mp_expected_start = datetime.strptime(mp_expected_start, "%Y-%m-%d")
            
            mp_expected_end = today - timedelta(days=2)
            
            # 生成期望日期
            expected_dates = set()
            curr = mp_expected_start
            while curr <= mp_expected_end:
                expected_dates.add(curr.strftime("%Y-%m-%d"))
                curr += timedelta(days=1)
            
            if not expected_dates:
                continue
            
            # 获取该客户的曲线数据
            curves = list(UNIFIED_LOAD_CURVE.find({"customer_id": customer_id}))
            
            # 1. 找出整天缺失的日期（完全没生成曲线记录的日期）
            actual_curve_dates = set(c.get("date") for c in curves)
            whole_day_missing = expected_dates - actual_curve_dates
            
            # 2. 找出有数据但MP缺失的日期（在missing_mps名单里的）
            mp_specific_missing = {} # mp_no -> set of dates
            for mp_no in mp_list:
                mp_specific_missing[mp_no] = set()

            for curve in curves:
                date = curve.get("date")
                if date not in expected_dates:
                    continue

                mp_load = curve.get("mp_load")
                
                # 如果 mp_load 不存在或为空（说明当天完全没有MP数据，但因为有电表数据所以生成了curve）
                if not mp_load:
                    for mp_no in mp_list:
                        mp_specific_missing[mp_no].add(date)
                    continue

                # 获取该日期明确缺失的计量点
                missing_mps_raw = mp_load.get("missing_mps", [])
                missing_mps = set(str(x) for x in missing_mps_raw)
                    
                for mp_no in mp_list:
                    if str(mp_no) in missing_mps:
                        mp_specific_missing[mp_no].add(date)
            
            # 生成缺失记录
            for mp_no in mp_list:
                # 总缺失 = 整天缺失 + 有记录但该MP缺失
                all_missing_dates = whole_day_missing | mp_specific_missing[mp_no]
                
                if all_missing_dates:
                    # 格式化日期列表
                    sorted_dates = sorted(all_missing_dates)
                    # 简化格式：月-日
                    formatted_dates = ", ".join(d[5:] for d in sorted_dates)  # 去掉年份 YYYY-
                    
                    rows.append({
                        "customer_name": customer_name,
                        "mp_no": mp_no,
                        "missing_dates": formatted_dates,
                        "missing_count": len(all_missing_dates)
                    })
        
        # 过滤掉没有缺失的记录（虽然上面逻辑已经保证了missing_dates才添加，但为了双重保险）
        rows = [r for r in rows if r["missing_count"] > 0]
        
        # 按客户名称和缺失天数排序
        rows.sort(key=lambda x: (x["customer_name"], -x["missing_count"], x["mp_no"]))
        
        # 生成Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "计量点缺失明细"
        
        # 表头
        headers = ["客户名称", "计量点号", "缺失日期", "缺失天数"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, row_data in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_data["customer_name"])
            ws.cell(row=row_idx, column=2, value=row_data["mp_no"])
            ws.cell(row=row_idx, column=3, value=row_data["missing_dates"])
            ws.cell(row=row_idx, column=4, value=row_data["missing_count"])
            
            for col in range(1, 5):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = Font(name='微软雅黑', size=10)
                cell.border = thin_border
                if col == 4:  # 缺失天数列右对齐
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 设置列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 12
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 添加筛选
        ws.auto_filter.ref = f"A1:D{len(rows) + 1}"
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"计量点缺失明细_{today.strftime('%Y%m%d')}.xlsx"
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
        
    except Exception as e:
        logger.error(f"导出计量点缺失明细失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
