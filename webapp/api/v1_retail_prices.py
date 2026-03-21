"""
零售结算价格定义 API
对应数据集：retail_settlement_prices
"""
import logging
import json
from datetime import datetime

from bson import json_util
from fastapi import APIRouter, File, UploadFile, Depends, HTTPException

from webapp.tools.mongo import DATABASE
from webapp.tools.security import get_current_active_user, User
from webapp.api.dependencies.authz import require_permission

logger = logging.getLogger(__name__)

router = APIRouter(tags=["零售结算价格"])

COLLECTION = DATABASE['retail_settlement_prices']

# 常规价格类型的中文名 → 英文键名映射（与套餐 reference_type 命名一致）
REGULAR_PRICE_KEY_MAP = {
    '中长期市场月度交易均价（不分时）': 'market_monthly_avg',
    '中长期市场年度交易均价（不分时）': 'market_annual_avg',
    '中长期市场交易均价（不分时）': 'market_avg',
    '中长期市场当月平均上网电价': 'market_monthly_on_grid',
    '售电公司月度结算加权价': 'retailer_monthly_settle_weighted',
    '售电公司月度交易均价（不分时）': 'retailer_monthly_avg',
    '售电公司年度交易均价（不分时）': 'retailer_annual_avg',
    '售电公司交易均价（不分时）': 'retailer_avg',
    '售电侧月度结算加权价': 'retailer_side_settle_weighted',
    '省内现货实时市场加权平均价': 'real_time_avg',
    '煤电容量电费折价': 'coal_capacity_discount',
    '发电侧火电年度中长期双边协商交易合同分月平段价': 'genside_annual_bilateral',
    '电网代理购电价格': 'grid_agency_price',
    '市场化用户中长期交易平段合同加权平均价': 'market_longterm_flat_avg',
}

# 分时价格 Excel 列顺序（第5列起）→ 字段名
PERIOD_PRICE_COLUMNS = [
    'upper_limit_price',
    'market_monthly_avg',
    'market_annual_avg',
    'market_avg',
    'market_monthly_on_grid',
    'retailer_monthly_avg',
    'retailer_annual_avg',
    'retailer_avg',
    'real_time_avg',
    'day_ahead_avg',
    'genside_annual_bilateral',
    'grid_agency_price',
]


def _parse_excel(content: bytes) -> dict:
    """解析 Excel 文件，返回 {regular_prices, period_prices}"""
    try:
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件解析失败: {str(e)}")

    sheet_names = wb.sheetnames
    regular_sheet = None
    period_sheet = None

    for name in sheet_names:
        if '常规价格' in name or '非分时' in name:
            regular_sheet = wb[name]
        elif '分时价格' in name:
            period_sheet = wb[name]

    if not regular_sheet or not period_sheet:
        raise HTTPException(
            status_code=400,
            detail=f"未找到所需工作表（需含'常规价格'和'分时价格'），当前工作表：{sheet_names}"
        )

    # 解析常规价格（第1行为标题，从第2行开始是数据）
    regular_prices = []
    month_str = None
    for row in regular_sheet.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        month_val, price_type, price_value, definition = row[0], row[1], row[2], row[3]
        if month_val and month_str is None:
            month_str = str(month_val).strip()
        price_type_str = str(price_type).strip() if price_type else ''
        key = REGULAR_PRICE_KEY_MAP.get(price_type_str, price_type_str)
        regular_prices.append({
            'price_type': price_type_str,
            'price_type_key': key,
            'value': float(price_value) if price_value is not None else None,
            'definition': str(definition).strip() if definition else '',
        })

    # 解析分时价格（找到标题行，其后为数据）
    period_prices = []
    data_start_row = None
    for i, row in enumerate(period_sheet.iter_rows(values_only=True), start=1):
        if row[0] == '月份' or row[0] == 'month':
            data_start_row = i + 1
            break

    if data_start_row is None:
        raise HTTPException(status_code=400, detail="分时价格工作表中未找到标题行（'月份'列）")

    for row in period_sheet.iter_rows(min_row=data_start_row, values_only=True):
        if not any(v is not None for v in row):
            continue
        month_val, period_no, period_type, float_ratio = row[0], row[1], row[2], row[3]
        if month_val and month_str is None:
            month_str = str(month_val).strip()

        period_data = {
            'period': int(period_no) if period_no is not None else None,
            'period_type': str(period_type).strip() if period_type else '',
            'float_ratio': float(float_ratio) if float_ratio is not None else None,
        }
        # 从第5列起映射价格字段
        for col_idx, field_name in enumerate(PERIOD_PRICE_COLUMNS):
            col_val = row[4 + col_idx] if (4 + col_idx) < len(row) else None
            period_data[field_name] = float(col_val) if col_val is not None else None

        period_prices.append(period_data)

    if not month_str:
        raise HTTPException(status_code=400, detail="无法从 Excel 文件中识别月份信息")

    return {
        'month': month_str,
        'regular_prices': regular_prices,
        'period_prices': period_prices,
    }


# ─────────────────────────────────────────────
# GET /prices/retail-settlement  获取已导入的月份列表
# ─────────────────────────────────────────────
@router.get("/prices/retail-settlement", summary="获取零售结算价格月份列表")
def list_retail_settlement_prices():
    """返回所有已导入的月份列表（按月份降序），以及最新月份的完整价格数据。"""
    try:
        docs = list(COLLECTION.find(
            {},
            {'_id': 1, 'month': 1, 'price_date_type': 1, 'imported_at': 1, 'imported_by': 1}
        ).sort('_id', -1))
        return json.loads(json_util.dumps({'months': docs}))
    except Exception as e:
        logger.error(f"list_retail_settlement_prices error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────
# GET /prices/retail-settlement/{month}  获取指定月份详细数据
# ─────────────────────────────────────────────
@router.get("/prices/retail-settlement/{month}", summary="获取指定月份零售结算价格")
def get_retail_settlement_price(month: str):
    """返回完整的常规价格和分时价格数据。"""
    try:
        doc = COLLECTION.find_one({'_id': month})
        if not doc:
            raise HTTPException(status_code=404, detail=f"月份 {month} 的价格数据不存在")
        return json.loads(json_util.dumps(doc))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"get_retail_settlement_price error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────
# POST /prices/retail-settlement/import  导入 Excel 文件
# ─────────────────────────────────────────────
@router.post("/prices/retail-settlement/import", summary="导入零售结算价格 Excel")
async def import_retail_settlement_prices(
    price_date_type: str = "regular",
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_monthly_manual_import:edit"))
):
    """
    上传并解析月度价格定义 Excel，存入 retail_settlement_prices 集合。
    
    Args:
        price_date_type: 价格适用日期类型，'regular' (常规/默认) 或 'holiday' (节假日/深谷)
        file: Excel 文件
    """
    if price_date_type not in ("regular", "holiday"):
        raise HTTPException(status_code=400, detail="无效的价格日期类型")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持 Excel 文件（.xlsx/.xls）")

    content = await file.read()
    parsed = _parse_excel(content)
    month = parsed['month']
    
    # 确定数据库 ID: 常规/默认使用 YYYY-MM，节假日使用 YYYY-MM-holiday
    doc_id = month if price_date_type == "regular" else f"{month}-holiday"

    now = datetime.now()
    doc = {
        '_id': doc_id,
        'month': month,
        'price_date_type': price_date_type,
        'imported_at': now,
        'imported_by': current_user.username,
        'regular_prices': parsed['regular_prices'],
        'period_prices': parsed['period_prices'],
    }

    # 同 ID 数据直接覆盖
    COLLECTION.replace_one({'_id': doc_id}, doc, upsert=True)

    logger.info(f"月份 {month} ({price_date_type}) 零售结算价格已导入，操作人：{current_user.username}")
    return {
        'status': 'success',
        'month': month,
        'regular_count': len(parsed['regular_prices']),
        'period_count': len(parsed['period_prices']),
    }


# ─────────────────────────────────────────────
# DELETE /prices/retail-settlement/{month}  删除指定月份
# ─────────────────────────────────────────────
@router.delete("/prices/retail-settlement/{month}", summary="删除指定月份零售结算价格")
def delete_retail_settlement_price(
    month: str,
    current_user: User = Depends(get_current_active_user),
    _ctx = Depends(require_permission("module:basic_monthly_manual_import:edit"))
):
    result = COLLECTION.delete_one({'_id': month})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail=f"月份 {month} 的价格数据不存在")
    return {'status': 'success', 'deleted_month': month}
