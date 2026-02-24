import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExportService:
    def __init__(self):
        pass

    def export_wholesale_to_excel(self, date_str: str, version_name: str, data: dict) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "批发侧结算详情"

        # 样式定义
        header_font = Font(name='微软雅黑', bold=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 定义颜色
        colors = {
            'blue': 'E3F2FD',    # 中长期
            'orange': 'FFF3E0',  # 日前
            'green': 'E8F5E9',   # 实时
            'red': 'FFEBEE',     # 电能量
            'purple': 'F3E5F5'   # 标准值
        }

        # 1. 概览信息
        ws.merge_cells('A1:P1')
        ws['A1'] = f"批发侧结算详情 - {date_str} ({version_name})"
        ws['A1'].font = Font(name='微软雅黑', bold=True, size=14)
        ws['A1'].alignment = center_align

        # 2. 第一级表头
        header1 = [
            ("时段", 1, 2),
            ("中长期合约电费", 3, 1),
            ("日前市场偏差", 3, 1),
            ("实时市场偏差", 3, 1),
            ("电能量", 2, 1),
            ("标准值", 4, 1)
        ]
        
        # 填充第一级表头并合并
        curr_col = 1
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 35

        # 辅助函数：应用样式
        def apply_style(cell_range, bg_color=None):
            res = ws[cell_range]
            # 如果是单单元格，包装成嵌套元组以统一迭代逻辑
            cells_to_style = ((res,),) if not isinstance(res, (tuple, list)) else res
            
            for row in cells_to_style:
                for cell in row:
                    cell.alignment = center_align
                    cell.border = thin_border
                    if bg_color:
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 时段
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.cell(row=2, column=1, value="时段")
        apply_style('A2:A3')
        curr_col = 2

        groups = [
            ("中长期合约电费", 3, colors['blue']),
            ("日前市场偏差", 3, colors['orange']),
            ("实时市场偏差", 3, colors['green']),
            ("电能量", 2, colors['red']),
            ("标准值", 4, colors['purple'])
        ]

        for text, span, color in groups:
            ws.merge_cells(start_row=2, start_column=curr_col, end_row=2, end_column=curr_col + span - 1)
            cell = ws.cell(row=2, column=curr_col, value=text)
            cell.font = header_font
            range_str = f"{ws.cell(row=2, column=curr_col).coordinate}:{ws.cell(row=2, column=curr_col + span - 1).coordinate}"
            apply_style(range_str, color)
            curr_col += span

        # 3. 第二级表头 (子列名称 + 公式)
        sub_headers = [
            ("合同电量\n①", colors['blue']),
            ("合同均价\n②", colors['blue']),
            ("差价电费\n③=①×(②-⑤)", colors['blue']),
            ("出清电量\n④", colors['orange']),
            ("市场均价\n⑤", colors['orange']),
            ("差价电费\n⑥=④×(⑤-⑧)", colors['orange']),
            ("实际用量\n⑦", colors['green']),
            ("市场均价\n⑧", colors['green']),
            ("全量电费\n⑨=⑦×⑧", colors['green']),
            ("电费合计\n⑩=③+⑥+⑨", colors['red']),
            ("结算均价\n⑪=⑩÷⑦", colors['red']),
            ("机制电量\n⑫", colors['purple']),
            ("签约比例\n⑬=(①+⑫)÷⑦", colors['purple']),
            ("电费合计\n⑭", colors['purple']),
            ("结算均价\n⑮=⑭÷⑦", colors['purple'])
        ]

        for i, (text, color) in enumerate(sub_headers):
            cell = ws.cell(row=3, column=i + 2, value=text)
            cell.font = Font(name='微软雅黑', size=9)
            apply_style(cell.coordinate, color)

        # 4. 填充数据
        details = data.get('wholesale_period_details', [])
        for idx, p in enumerate(details):
            row_idx = idx + 4
            # A: 时段
            ws.cell(row=row_idx, column=1, value=p.get('period'))
            
            # B: 合同电量 ①
            ws.cell(row=row_idx, column=2, value=p.get('contract', {}).get('volume', 0))
            # C: 合同均价 ②
            ws.cell(row=row_idx, column=3, value=p.get('contract', {}).get('price', 0))
            # D: 差价电费 ③ = B * (C - F)  [C is Price2, F is Price5]
            ws.cell(row=row_idx, column=4, value=f"=B{row_idx}*(C{row_idx}-F{row_idx})")
            
            # E: 出清电量 ④
            ws.cell(row=row_idx, column=5, value=p.get('day_ahead', {}).get('volume', 0))
            # F: 市场均价 ⑤
            ws.cell(row=row_idx, column=6, value=p.get('day_ahead', {}).get('price', 0))
            # G: 差价电费 ⑥ = E * (F - I) [I is Price8]
            ws.cell(row=row_idx, column=7, value=f"=E{row_idx}*(F{row_idx}-I{row_idx})")
            
            # H: 实际用量 ⑦
            ws.cell(row=row_idx, column=8, value=p.get('real_time', {}).get('volume', 0))
            # I: 市场均价 ⑧
            ws.cell(row=row_idx, column=9, value=p.get('real_time', {}).get('price', 0))
            # J: 全量电费 ⑨ = H * I
            ws.cell(row=row_idx, column=10, value=f"=H{row_idx}*I{row_idx}")
            
            # K: 电费合计 ⑩ = D + G + J
            ws.cell(row=row_idx, column=11, value=f"=D{row_idx}+G{row_idx}+J{row_idx}")
            # L: 结算均价 ⑪ = K / H
            ws.cell(row=row_idx, column=12, value=f"=IF(H{row_idx}=0, 0, K{row_idx}/H{row_idx})")
            
            # M: 机制电量 ⑫
            ws.cell(row=row_idx, column=13, value=p.get('mechanism_volume', 0))
            # N: 签约比例 ⑬ = (B + M) / H
            ws.cell(row=row_idx, column=14, value=f"=IF(H{row_idx}=0, 0, (B{row_idx}+M{row_idx})/H{row_idx})")
            # O: 标准值电费 ⑭ (从后端获取值，因为逻辑复杂)
            ws.cell(row=row_idx, column=15, value=p.get('standard_value_cost', 0))
            # P: 结算均价 ⑮ = O / H
            ws.cell(row=row_idx, column=16, value=f"=IF(H{row_idx}=0, 0, O{row_idx}/H{row_idx})")

            # 设置数字格式和边框
            for col in range(1, 17):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                if col > 1:
                    cell.number_format = '0.00' if col in [3, 4, 6, 7, 9, 10, 11, 12, 15, 16] else '0.000'
                    if col == 14: # 签约比例
                        cell.number_format = '0.0%'

        # 5. 合计行
        total_row_idx = 4 + len(details)
        ws.cell(row=total_row_idx, column=1, value="合计").font = header_font
        for col in [2, 4, 5, 7, 8, 10, 11, 13, 15]: # 电量和费用的求和
            col_letter = get_column_letter(col)
            ws.cell(row=total_row_idx, column=col, value=f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})")
        
        # 合计行的均价/比例公式
        # K, L, N, P
        ws.cell(row=total_row_idx, column=12, value=f"=IF(H{total_row_idx}=0, 0, K{total_row_idx}/H{total_row_idx})") # ⑪
        ws.cell(row=total_row_idx, column=14, value=f"=IF(H{total_row_idx}=0, 0, (B{total_row_idx}+M{total_row_idx})/H{total_row_idx})") # ⑬
        ws.cell(row=total_row_idx, column=16, value=f"=IF(H{total_row_idx}=0, 0, O{total_row_idx}/H{total_row_idx})") # ⑮

        # 合计行样式
        for col in range(1, 17):
            cell = ws.cell(row=total_row_idx, column=col)
            cell.font = header_font
            cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
            cell.border = thin_border
            if col > 1:
                col_letter = get_column_letter(col)
                if col == 14: cell.number_format = '0.0%'
                else: cell.number_format = '0.00'

        # 列宽微调
        for col in range(1, 17):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 12

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
